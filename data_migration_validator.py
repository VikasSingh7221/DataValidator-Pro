
import os
import hashlib
import pandas as pd
import traceback
import warnings
import json
import argparse
import glob
import sys
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
import logging

warnings.filterwarnings("ignore")

# ── Database drivers ──────────────────────────────────────────────────────────
try:
    import psycopg2
except ImportError:
    os.system("pip install psycopg2-binary --break-system-packages")
    import psycopg2

try:
    import snowflake.connector
except ImportError:
    os.system("pip install snowflake-connector-python --break-system-packages")
    import snowflake.connector

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    os.system("pip install openpyxl --break-system-packages")
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler("row_hash_validation.log"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# ── Data classes ──────────────────────────────────────────────────────────────

@dataclass
class ValidationResult:
    """One validation result record (one row comparison)."""
    table_name: str
    validation_type: str        # ROW_HASH_<i>
    status: str                 # PASS | FAIL | WARNING
    source_value: Any           # RS row MD5
    target_value: Any           # SF row MD5
    difference: Any = None
    details: str = ""
    timestamp: str = datetime.now().isoformat()


@dataclass
class TableValidation:
    """Per-table summary."""
    table_name: str
    total_checks: int
    passed_checks: int
    failed_checks: int
    warning_checks: int
    pass_rate: float
    row_count_source: int
    row_count_target: int
    row_count_diff_pct: float
    sampled_rows: int
    duration_seconds: float
    status: str
    details: str = ""


# ── Redshift connection ───────────────────────────────────────────────────────

class RedshiftConnection:
    def __init__(self, host: str, port: int, database: str, user: str, password: str):
        self.host = host
        self.port = port
        self.database = database
        self.user = user
        self.password = password
        self.conn = None

    def connect(self) -> bool:
        try:
            self.conn = psycopg2.connect(
                host=self.host, port=self.port, database=self.database,
                user=self.user, password=self.password, connect_timeout=10,
            )
            logger.info(f"■ Connected to Redshift: {self.host}")
            return True
        except Exception as e:
            logger.error(f"■ Redshift connection failed: {e}")
            return False

    def disconnect(self):
        if self.conn:
            self.conn.close()

    def execute_query(self, query: str) -> pd.DataFrame:
        try:
            return pd.read_sql_query(query, self.conn)
        except Exception as e:
            logger.error(f"Redshift query failed: {e}")
            raise

    def get_table_list(self, schema: str = "public", tables: Optional[List[str]] = None) -> List[str]:
        query = f"""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = '{schema}' AND table_type = 'BASE TABLE'
        """
        if tables:
            tables_list = ",".join([f"'{t.upper()}'" for t in tables])
            query += f" AND UPPER(table_name) IN ({tables_list})"
        query += "\n        ORDER BY table_name"
        df = self.execute_query(query)
        return df["table_name"].str.lower().tolist()

    def get_row_count(self, table_name: str, schema: str = "public") -> int:
        try:
            df = self.execute_query(f"SELECT COUNT(*) as cnt FROM {schema}.{table_name}")
            return int(df["cnt"].iloc[0])
        except Exception as e:
            logger.warning(f"Could not get row count for {table_name}: {e}")
            return 0

    def get_sample_data(self, table_name: str, sample_size: int = 100,
                        schema: str = "public", pk: str = None) -> pd.DataFrame:
        order_clause = f"ORDER BY {pk}" if pk else "ORDER BY 1"
        query = f"SELECT * FROM {schema}.{table_name} {order_clause} LIMIT {sample_size}"
        return self.execute_query(query)


# ── Snowflake connection ──────────────────────────────────────────────────────

class SnowflakeConnection:
    def __init__(self, account: str, user: str, password: str,
                 warehouse: str, database: str, mfa_code: Optional[int] = None, sso: bool = False):
        self.account = account
        self.user = user
        self.password = password
        self.warehouse = warehouse
        self.database = database
        self.mfa_code = mfa_code
        self.sso = sso
        self.conn = None

    def connect(self) -> bool:
        try:
            connect_args = {
                "account": self.account,
                "user": self.user,
                "warehouse": self.warehouse,
                "database": self.database,
            }
            if self.sso:
                connect_args["authenticator"] = "externalbrowser"
            else:
                connect_args["password"] = self.password
                if self.mfa_code is not None:
                    connect_args["passcode"] = self.mfa_code

            self.conn = snowflake.connector.connect(**connect_args)
            logger.info(f"■ Connected to Snowflake: {self.account}")
            return True
        except Exception as e:
            logger.error(f"■ Snowflake connection failed: {e}")
            return False

    def disconnect(self):
        if self.conn:
            self.conn.close()

    def execute_query(self, query: str) -> pd.DataFrame:
        try:
            cursor = self.conn.cursor()
            cursor.execute(query)
            df = cursor.fetch_pandas_all()
            cursor.close()
            return df
        except Exception as e:
            logger.error(f"Snowflake query failed: {e}")
            raise

    def get_table_list(self, schema: str = "tenant1", tables: Optional[List[str]] = None) -> List[str]:
        query = f"""
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = '{schema.upper()}'
        """
        if tables:
            tables_list = ",".join([f"'{t.lower()}'" for t in tables])
            query += f" AND lower(table_name) IN ({tables_list})"
        query += "\n        ORDER BY table_name"
        df = self.execute_query(query)
        return df["TABLE_NAME"].str.lower().tolist()

    def get_row_count(self, table_name: str, schema: str = "PUBLIC") -> int:
        try:
            df = self.execute_query(f"SELECT COUNT(*) as cnt FROM {schema}.{table_name}")
            return int(df["CNT"].iloc[0])
        except Exception as e:
            logger.warning(f"Could not get row count for {table_name}: {e}")
            return 0

    def get_sample_data(self, table_name: str, sample_size: int = 100,
                        schema: str = "PUBLIC", pk: str = None) -> pd.DataFrame:
        order_clause = f"ORDER BY {pk.upper()}" if pk else "ORDER BY 1"
        query = f"SELECT * FROM {schema.upper()}.{table_name.upper()} {order_clause} LIMIT {sample_size}"
        return self.execute_query(query)


# ── Validator ─────────────────────────────────────────────────────────────────

class DataValidator:
    """Row-hash-only validation: the ONLY check is per-row MD5 comparison."""

    def __init__(self, redshift_conn: RedshiftConnection, snowflake_conn: SnowflakeConnection,
                 sample_size: int = 100, table_pks: dict = None, tenant_name: str = "tenant",
                 group_name: str = "standard", query_name: str = None):
        self.rs = redshift_conn
        self.sf = snowflake_conn
        self.sample_size = sample_size
        self.table_pks = table_pks or {}
        self.tenant_name = tenant_name
        self.group_name = group_name
        self.query_name = query_name
        self.validation_results: List[ValidationResult] = []
        self.table_summaries: List[TableValidation] = []

        ist_tz = timezone(timedelta(hours=5, minutes=30))
        self.run_time = datetime.now(ist_tz).strftime("%Y%m%d_%H%M%S")
        if self.query_name:
            self.report_folder = os.path.join(self.tenant_name, f"{self.tenant_name}_{self.query_name}_comparison_{self.run_time}")
        else:
            self.report_folder = os.path.join(self.tenant_name, f"{self.tenant_name}_{self.group_name}_table_wise_comparison_{self.run_time}")

    # ── public ────────────────────────────────────────────────────────────────

    def validate_all_tables(self, rs_schema: str = "public", sf_schema: str = "PUBLIC",
                            tables: Optional[List[str]] = None) -> Dict[str, TableValidation]:
        if tables is None:
            target_tables = list(self.table_pks.keys()) if self.table_pks else None
            rs_tables = set(self.rs.get_table_list(rs_schema, tables=target_tables))
            sf_tables = set(self.sf.get_table_list(sf_schema, tables=target_tables))
            tables = list(rs_tables & sf_tables)

            missing_in_sf = rs_tables - sf_tables
            missing_in_rs = sf_tables - rs_tables
            if missing_in_sf:
                logger.warning(f"Tables in Redshift but NOT in Snowflake: {missing_in_sf}")
            if missing_in_rs:
                logger.warning(f"Tables in Snowflake but NOT in Redshift: {missing_in_rs}")

        logger.info(f"Validating {len(tables)} tables...")

        results = {}
        for idx, table in enumerate(tables, 1):
            logger.info(f"\n[{idx}/{len(tables)}] Validating {table}...")
            try:
                table_result = self.validate_table(table, rs_schema, sf_schema)
                results[table] = table_result
                self.table_summaries.append(table_result)
            except Exception as e:
                logger.error(f"Error validating {table}: {e}")
                traceback.print_exc()

        return results

    def validate_table(self, table_name: str,
                       rs_schema: str = "public", sf_schema: str = "PUBLIC") -> TableValidation:
        """Single table: fetch sample data, compare row-by-row via MD5."""
        pk = self.table_pks.get(table_name.lower())
        start_time = datetime.now()

        checks_passed = 0
        checks_failed = 0
        checks_warning = 0
        details = []

        # Fetch row counts (informational only — not a pass/fail check)
        rs_count = self.rs.get_row_count(table_name, rs_schema)
        sf_count = self.sf.get_row_count(table_name, sf_schema)
        details.append(f"Row counts: RS={rs_count}, SF={sf_count}")

        # ── Row-hash validation ───────────────────────────────────────────────
        if rs_count > 0 and sf_count > 0:
            try:
                rs_sample = self.rs.get_sample_data(table_name, self.sample_size, rs_schema, pk=pk)
                sf_sample = self.sf.get_sample_data(table_name, self.sample_size, sf_schema, pk=pk)

                row_results = self._row_hash_validation(table_name, rs_sample, sf_sample)

                row_mismatches = 0
                for result in row_results:
                    self.validation_results.append(result)
                    if result.status == "PASS":
                        checks_passed += 1
                    elif result.status == "FAIL":
                        checks_failed += 1
                        row_mismatches += 1
                    else:
                        checks_warning += 1

                total_compared = len(row_results)
                if row_mismatches == 0:
                    details.append(f"Row-hash: All {total_compared} rows matched")
                else:
                    details.append(f"Row-hash: {row_mismatches}/{total_compared} rows mismatched")

                # ── TABLE_WISE folder ─────────────────────────────────────────
                folder_path = os.path.join(self.report_folder, table_name)
                os.makedirs(folder_path, exist_ok=True)

                rs_sample.to_csv(
                    os.path.join(folder_path, f"{table_name}_rs_sample.csv"), index=False)
                sf_sample.to_csv(
                    os.path.join(folder_path, f"{table_name}_sf_sample.csv"), index=False)

                # Row-level comparison CSV
                comp_rows = []
                for res in row_results:
                    comp_rows.append({
                        "Row_Index": res.validation_type.replace("ROW_HASH_", ""),
                        "Status": res.status,
                        "RS_Row_MD5": res.source_value,
                        "SF_Row_MD5": res.target_value,
                        "Details": res.details,
                    })
                pd.DataFrame(comp_rows).to_csv(
                    os.path.join(folder_path, f"{table_name}_comparison.csv"), index=False)

                # Coloured side-by-side mismatch Excel
                self._write_mismatch_excel(
                    folder_path, table_name, rs_sample.copy(), sf_sample.copy()
                )

            except Exception as e:
                logger.warning(f"Row-hash validation failed for {table_name}: {e}")
                details.append(f"Validation skipped: {e}")
        else:
            details.append("Skipped: one or both tables have 0 rows")

        # ── Summary ───────────────────────────────────────────────────────────
        total_checks = checks_passed + checks_failed + checks_warning
        pass_rate = (checks_passed / total_checks * 100) if total_checks > 0 else 0.0

        if checks_failed == 0:
            status = "PASS" if checks_warning == 0 else "WARNING"
        else:
            status = "FAIL"

        duration = (datetime.now() - start_time).total_seconds()

        pct_diff = round(abs(rs_count - sf_count) / (rs_count if rs_count > 0 else 1) * 100, 2)

        return TableValidation(
            table_name=table_name,
            total_checks=total_checks,
            passed_checks=checks_passed,
            failed_checks=checks_failed,
            warning_checks=checks_warning,
            pass_rate=pass_rate,
            row_count_source=rs_count,
            row_count_target=sf_count,
            row_count_diff_pct=pct_diff,
            sampled_rows=min(self.sample_size, rs_count, sf_count),
            duration_seconds=duration,
            status=status,
            details=" | ".join(details),
        )

    def validate_custom_query(self, rs_col_df: pd.DataFrame, sf_col_df: pd.DataFrame) -> TableValidation:
        """Evaluate explicitly provided samples."""
        table_name = self.query_name if self.query_name else "Custom_Query"
        start_time = datetime.now()

        checks_passed = 0
        checks_failed = 0
        checks_warning = 0
        details = []

        rs_count = len(rs_col_df)
        sf_count = len(sf_col_df)
        details.append(f"Row counts: RS={rs_count}, SF={sf_count}")

        if rs_count > 0 and sf_count > 0:
            try:
                row_results = self._row_hash_validation(table_name, rs_col_df, sf_col_df)

                row_mismatches = 0
                for result in row_results:
                    self.validation_results.append(result)
                    if result.status == "PASS":
                        checks_passed += 1
                    elif result.status == "FAIL":
                        checks_failed += 1
                        row_mismatches += 1
                    else:
                        checks_warning += 1

                total_compared = len(row_results)
                if row_mismatches == 0:
                    details.append(f"Row-hash: All {total_compared} rows matched")
                else:
                    details.append(f"Row-hash: {row_mismatches}/{total_compared} rows mismatched")

                folder_path = os.path.join(self.report_folder, table_name)
                os.makedirs(folder_path, exist_ok=True)

                rs_col_df.to_csv(os.path.join(folder_path, f"{table_name}_rs_sample.csv"), index=False)
                sf_col_df.to_csv(os.path.join(folder_path, f"{table_name}_sf_sample.csv"), index=False)

                comp_rows = []
                for res in row_results:
                    comp_rows.append({
                        "Row_Index": res.validation_type.replace("ROW_HASH_", ""),
                        "Status": res.status,
                        "RS_Row_MD5": res.source_value,
                        "SF_Row_MD5": res.target_value,
                        "Details": res.details,
                    })
                pd.DataFrame(comp_rows).to_csv(os.path.join(folder_path, f"{table_name}_comparison.csv"), index=False)

                self._write_mismatch_excel(folder_path, table_name, rs_col_df.copy(), sf_col_df.copy())

            except Exception as e:
                logger.warning(f"Row-hash validation failed for {table_name}: {e}")
                details.append(f"Validation skipped: {e}")
        else:
            details.append("Skipped: one or both queries returned 0 rows")

        total_checks = checks_passed + checks_failed + checks_warning
        pass_rate = (checks_passed / total_checks * 100) if total_checks > 0 else 0.0

        if checks_failed == 0:
            status = "PASS" if checks_warning == 0 else "WARNING"
        else:
            status = "FAIL"

        duration = (datetime.now() - start_time).total_seconds()

        pct_diff = round(abs(rs_count - sf_count) / (rs_count if rs_count > 0 else 1) * 100, 2)

        return TableValidation(
            table_name=table_name,
            total_checks=total_checks,
            passed_checks=checks_passed,
            failed_checks=checks_failed,
            warning_checks=checks_warning,
            pass_rate=pass_rate,
            row_count_source=rs_count,
            row_count_target=sf_count,
            row_count_diff_pct=pct_diff,
            sampled_rows=min(rs_count, sf_count),
            duration_seconds=duration,
            status=status,
            details=" | ".join(details),
        )

    # ── private ───────────────────────────────────────────────────────────────

    @staticmethod
    def _norm_val(v) -> str:
        """Normalise a single cell value to a comparable string."""
        if isinstance(v, float) and pd.notna(v) and v.is_integer():
            v = int(v)
            
        val_str = str(v).strip()
        if val_str.endswith('.0'):
            try:
                if float(val_str).is_integer():
                    val_str = val_str[:-2]
            except ValueError:
                pass
                
        return (
            val_str
            .replace("nan", "")
            .replace("None", "")
            .replace("NaT", "")
            .replace("<NA>", "")
        )

    def _row_hash(self, row: pd.Series) -> str:
        """Return a 12-char MD5 hex digest for one normalised row."""
        concatenated = "|".join(self._norm_val(v) for v in row)
        return hashlib.md5(concatenated.encode("utf-8")).hexdigest()[:12]

    def _row_hash_validation(self, table_name: str,
                             rs_df: pd.DataFrame,
                             sf_df: pd.DataFrame) -> List[ValidationResult]:
        """Compare each sampled row via MD5. Returns one result per row."""
        rs_df = rs_df.copy()
        sf_df = sf_df.copy()
        rs_df.columns = rs_df.columns.str.lower()
        sf_df.columns = sf_df.columns.str.lower()

        common_cols = sorted(set(rs_df.columns) & set(sf_df.columns))

        if not common_cols:
            return [ValidationResult(
                table_name=table_name,
                validation_type="ROW_HASH_DATA",
                status="WARNING",
                source_value="N/A",
                target_value="N/A",
                details="No common columns found between RS and SF",
            )]

        min_len = min(len(rs_df), len(sf_df))
        rs_sub = rs_df[common_cols].head(min_len).reset_index(drop=True)
        sf_sub = sf_df[common_cols].head(min_len).reset_index(drop=True)

        results = []
        for i in range(min_len):
            rs_hash = self._row_hash(rs_sub.iloc[i])
            sf_hash = self._row_hash(sf_sub.iloc[i])

            if rs_hash == sf_hash:
                status = "PASS"
                detail = "Row matches"
            else:
                status = "FAIL"
                # Surface which columns differ for easy debugging
                diff_cols = [
                    col for col in common_cols
                    if self._norm_val(rs_sub.iloc[i][col]) != self._norm_val(sf_sub.iloc[i][col])
                ]
                detail = f"Mismatch in columns: {diff_cols}"

            results.append(ValidationResult(
                table_name=table_name,
                validation_type=f"ROW_HASH_{i}",
                status=status,
                source_value=rs_hash,
                target_value=sf_hash,
                details=detail,
            ))

        return results

    def _write_mismatch_excel(self, folder_path: str, table_name: str,
                              rs_sample: pd.DataFrame, sf_sample: pd.DataFrame):
        """Write a coloured side-by-side diff Excel file for mismatched cells."""
        try:
            from openpyxl import load_workbook

            rs_sample.columns = rs_sample.columns.str.lower()
            sf_sample.columns = sf_sample.columns.str.lower()
            
            common_cols = sorted(set(rs_sample.columns) & set(sf_sample.columns))
            min_len = min(len(rs_sample), len(sf_sample))

            if min_len == 0 or not common_cols:
                return

            rs_sub = rs_sample[common_cols].head(min_len).reset_index(drop=True).copy()
            sf_sub = sf_sample[common_cols].head(min_len).reset_index(drop=True).copy()

            # Build display DataFrame — every cell shows "rs_val = sf_val"
            comparison_df = pd.DataFrame(index=range(min_len), columns=common_cols)
            color_mask = pd.DataFrame(False, index=range(min_len), columns=common_cols)

            for col in common_cols:
                rs_norm = rs_sub[col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).replace(
                    ["nan", "None", "NaT", "<NA>"], "")
                sf_norm = sf_sub[col].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).replace(
                    ["nan", "None", "NaT", "<NA>"], "")
                mismatch_mask = rs_norm != sf_norm
                color_mask[col] = mismatch_mask

                # Strip XML-illegal characters that cause openpyxl to crash
                illegal_chars = r'[\x00-\x08\x0b\x0c\x0e-\x1f]'
                rs_strs = rs_sub[col].astype(str).str.replace(illegal_chars, '', regex=True)
                sf_strs = sf_sub[col].astype(str).str.replace(illegal_chars, '', regex=True)

                # All cells: "RS:[ val ]  =  SF:[ val ]"  (same format as mismatch, green bg)
                comparison_df[col] = "RS:[ " + rs_strs + " ]  =  SF:[ " + sf_strs + " ]"

                # Overwrite mismatched cells: "RS:[ val ]  v  SF:[ val ]"  (red bg)
                if mismatch_mask.any():
                    diff_strings = (
                        "RS:[ " + rs_strs + " ]  v  SF:[ " + sf_strs + " ]"
                    )
                    comparison_df.loc[mismatch_mask, col] = diff_strings[mismatch_mask]

            mismatch_path = os.path.join(folder_path, f"{table_name}_row_mismatches.xlsx")
            comparison_df.to_excel(mismatch_path, index=False)

            wb = load_workbook(mismatch_path)
            ws = wb.active

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill   = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

            for col_i, col in enumerate(common_cols):
                for row_j in range(min_len):
                    cell = ws.cell(row=row_j + 2, column=col_i + 1)
                    cell.fill = red_fill if color_mask[col].iloc[row_j] else green_fill

            wb.save(mismatch_path)

        except Exception as e:
            logger.warning(f"Could not write mismatch Excel for {table_name}: {e}")


# ── Report Generator ──────────────────────────────────────────────────────────

class ReportGenerator:
    """Generates the same multi-sheet Excel report structure."""

    def __init__(self, validator: DataValidator):
        self.validator = validator

    def generate_excel_report(self):
        if self.validator.query_name:
            filename = f"{self.validator.tenant_name}_{self.validator.query_name}_comparison_{self.validator.run_time}.xlsx"
        else:
            filename = f"{self.validator.tenant_name}_{self.validator.group_name}_table_wise_comparison_{self.validator.run_time}.xlsx"
        filepath = os.path.join(self.validator.report_folder, filename)
        logger.info(f"Generating Excel report: {filepath}")

        os.makedirs(self.validator.report_folder, exist_ok=True)
        wb = Workbook()
        wb.remove(wb.active)

        self._create_summary_sheet(wb)
        self._create_details_sheet(wb)

        wb.save(filepath)
        logger.info(f"■ Excel report saved: {filepath}")
        return filepath

    # ── sheets ────────────────────────────────────────────────────────────────

    def _header_style(self, ws):
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _status_color(self, status: str) -> str:
        return {"PASS": "70AD47", "WARNING": "FFC7CE", "FAIL": "FF0000"}.get(status, "FFFFFF")

    def _create_summary_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Summary", 0)
        ws.append([
            "Table Name", "Status", "Total Row Checks", "Passed", "Failed",
            "Warnings", "Pass Rate %", "Source Rows", "Target Rows", "Row Diff %",
            "Sampled Rows", "Duration (s)",
        ])
        self._header_style(ws)

        for s in self.validator.table_summaries:
            ws.append([
                s.table_name, s.status, s.total_checks, s.passed_checks,
                s.failed_checks, s.warning_checks, f"{s.pass_rate:.1f}%",
                s.row_count_source, s.row_count_target, f"{s.row_count_diff_pct:.2f}%",
                s.sampled_rows,
                f"{s.duration_seconds:.2f}",
            ])
            ws.cell(row=ws.max_row, column=2).fill = PatternFill(
                start_color=self._status_color(s.status),
                end_color=self._status_color(s.status),
                fill_type="solid",
            )

        ws.column_dimensions["A"].width = 40
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
            ws.column_dimensions[col].width = 16

    def _create_details_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Detailed Results")
        ws.append([
            "Table", "Row Index", "Status",
            "RS Row MD5", "SF Row MD5", "Details",
        ])
        self._header_style(ws)

        for r in self.validator.validation_results:
            row_idx = r.validation_type.replace("ROW_HASH_", "")
            ws.append([
                r.table_name, row_idx, r.status,
                str(r.source_value), str(r.target_value),
                r.details[:200],
            ])
            ws.cell(row=ws.max_row, column=3).fill = PatternFill(
                start_color=self._status_color(r.status),
                end_color=self._status_color(r.status),
                fill_type="solid",
            )

        ws.column_dimensions["A"].width = 40
        for col in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 20

    def _create_failures_sheet(self, wb: Workbook):
        failed   = [r for r in self.validator.validation_results if r.status == "FAIL"]
        warnings = [r for r in self.validator.validation_results if r.status == "WARNING"]

        ws = wb.create_sheet("Failures & Warnings")

        ws.append(["FAILURES:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append(["Table", "Row Index", "RS MD5", "SF MD5", "Details"])
        self._header_style(ws)

        for r in failed:
            ws.append([
                r.table_name,
                r.validation_type.replace("ROW_HASH_", ""),
                str(r.source_value),
                str(r.target_value),
                r.details[:200],
            ])

        ws.append([])
        ws.append(["WARNINGS:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
        ws.append(["Table", "Row Index", "RS MD5", "SF MD5", "Details"])
        for r in warnings:
            ws.append([
                r.table_name,
                r.validation_type.replace("ROW_HASH_", ""),
                str(r.source_value),
                str(r.target_value),
                r.details[:200],
            ])

        ws.column_dimensions["A"].width = 40
        for col in ["B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 20

    def _create_statistics_sheet(self, wb: Workbook):
        ws = wb.create_sheet("Statistics")

        summ = self.validator.table_summaries
        total_tables   = len(summ)
        passed_tables  = sum(1 for t in summ if t.status == "PASS")
        failed_tables  = sum(1 for t in summ if t.status == "FAIL")
        warning_tables = sum(1 for t in summ if t.status == "WARNING")

        total_checks  = sum(t.total_checks  for t in summ)
        passed_checks = sum(t.passed_checks for t in summ)
        failed_checks = sum(t.failed_checks for t in summ)

        def pct(n, d): return f"{n} ({n/d*100:.1f}%)" if d > 0 else f"{n} (N/A)"

        stats = [
            ["ROW-HASH VALIDATION SUMMARY", ""],
            ["Validation Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["", ""],
            ["TABLE STATISTICS", ""],
            ["Total Tables Validated", total_tables],
            ["Tables Passed",          pct(passed_tables,  total_tables)],
            ["Tables Failed",          pct(failed_tables,  total_tables)],
            ["Tables with Warnings",   pct(warning_tables, total_tables)],
            ["", ""],
            ["ROW-HASH CHECK RESULTS", ""],
            ["Total Row Checks",  total_checks],
            ["Rows Matched",      pct(passed_checks, total_checks)],
            ["Rows Mismatched",   pct(failed_checks, total_checks)],
        ]

        for row in stats:
            ws.append(row)

        for row_cells in ws.iter_rows(min_row=1, max_row=len(stats)):
            val = row_cells[0].value
            if val and isinstance(val, str) and val.isupper():
                for cell in row_cells:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Row Hash Validator")
    parser.add_argument("-t", "--tenant", required=True, help="Tenant Name (e.g., tenant1)")
    parser.add_argument("-q", "--query", action="store_true", help="Run via queries in query.py")
    parser.add_argument("--sso", action="store_true", help="Use Snowflake SSO (externalbrowser)")
    parser.add_argument("--mfa", action="store_true", help="Use Snowflake MFA")
    parser.add_argument("-g", "--group", default="standard", help="Table mapping group (e.g., standard, claims, toc). Default: standard")
    args = parser.parse_args()

    tenant_name = args.tenant

    logger.info("=" * 60)
    logger.info(f"Redshift → Snowflake  |  Row-Hash Validator [{tenant_name}]")
    logger.info("=" * 60)

    # ── Load credentials ──────────────────────────────────────────────────────
    try:
        with open("config/credentials.json", "r") as f:
            creds = json.load(f)
    except Exception as e:
        logger.error(f"Failed to load config/credentials.json: {e}")
        return

    rs_cred = creds.get("redshift", {})
    sf_cred = creds.get("snowflake", {})

    rs_host = rs_cred.get("host", "localhost")
    rs_port = rs_cred.get("port", 5439)
    rs_database = rs_cred.get("database", "")
    rs_user = rs_cred.get("user", "")
    rs_password = rs_cred.get("password", "")
    rs_schema = tenant_name

    sf_account = sf_cred.get("account", "")
    sf_user = sf_cred.get("user", "")
    sf_password = sf_cred.get("password", "")
    sf_warehouse = sf_cred.get("warehouse", "")
    sf_database = sf_cred.get("database", "")
    sf_schema = tenant_name

    # ── Table PKs ─────────────────────────────────────────────────────────────
    table_pks = {}
    if not args.query:
        if os.path.isdir("tables"):
            # Try {group}_aggregation.json first, then {group}.json
            target_json = f"tables/{args.group}_aggregation.json"
            if not os.path.isfile(target_json):
                target_json = f"tables/{args.group}.json"

            if os.path.isfile(target_json):
                try:
                    with open(target_json, "r") as f:
                        lines = f.readlines()
                    
                    import re
                    cleaned_lines = []
                    for line in lines:
                        stripped = line.strip()
                        # Allow // or # for comments
                        if not (stripped.startswith('//') or stripped.startswith('#')):
                            cleaned_lines.append(line)
                            
                    cleaned_json = "".join(cleaned_lines)
                    # Handle trailing commas before closing braces/brackets
                    cleaned_json = re.sub(r',\s*([}\]])', r'\1', cleaned_json)
                    
                    data = json.loads(cleaned_json)
                    # Lowercase all keys for case-insensitive lookup
                    data = {k.lower(): v for k, v in data.items()}
                    table_pks.update(data)
                except Exception as e:
                    logger.warning(f"Could not load PK dict from {target_json}: {e}")
            else:
                # If still not found, list available groups to help the user
                available_files = glob.glob("tables/*.json")
                available_groups = []
                for f in available_files:
                    base = os.path.basename(f)
                    if base.endswith("_aggregation.json"):
                        available_groups.append(base.replace("_aggregation.json", ""))
                    elif base.endswith(".json"):
                        available_groups.append(base.replace(".json", ""))
                
                logger.error(f"Mapping file not found for group '{args.group}'.")
                logger.info(f"Looked for: tables/{args.group}_aggregation.json or tables/{args.group}.json")
                if available_groups:
                    logger.info(f"Available groups: {', '.join(sorted(set(available_groups)))}")
                return
        else:
            logger.warning("'tables' directory not found. Please create it for PK validation mapping.")

    # ── Connections ───────────────────────────────────────────────────────────
    logger.info("\nEstablishing connections...")

    rs_conn = RedshiftConnection(rs_host, rs_port, rs_database, rs_user, rs_password)
    if not rs_conn.connect():
        logger.error("Failed to connect to Redshift. Exiting.")
        return

    mfa_code = None
    if args.mfa and not args.sso:
        mfa_input = input("Enter MFA code: ").strip()
        if mfa_input:
            mfa_code = int(mfa_input)

    sf_conn = SnowflakeConnection(
        sf_account, sf_user, sf_password, sf_warehouse, sf_database, 
        mfa_code=mfa_code, sso=args.sso
    )
    if not sf_conn.connect():
        logger.error("Failed to connect to Snowflake. Exiting.")
        rs_conn.disconnect()
        return

    # ── Sample size ───────────────────────────────────────────────────────────
    sample_size = 100
    if not args.query:
        size_input = input("  Sample size for row-hash validation (100-500, default 100): ").strip()
        if size_input:
            sample_size = min(max(int(size_input), 100), 500)

    # Check for query_name before validator init
    query_name = None
    if args.query:
        try:
            import query
            query_name = getattr(query, 'query_name', 'Custom_Query')
        except ImportError:
            logger.error("Could not import query.py. Please create it with rs_query and sf_query variables.")
            return

    # ── Run validation ────────────────────────────────────────────────────────
    try:
        validator = DataValidator(
            rs_conn, sf_conn, sample_size=sample_size, table_pks=table_pks, 
            tenant_name=tenant_name, group_name=args.group, query_name=query_name
        )
        logger.info(f"\nStarting row-hash validation  |  sample_size={sample_size}")

        results = {}
        if args.query:
            logger.info(f"Running custom queries from query.py (name: {query_name})...")
            try:
                rs_df = rs_conn.execute_query(query.rs_query)
                sf_df = sf_conn.execute_query(query.sf_query)
                table_result = validator.validate_custom_query(rs_df, sf_df)
                results[query_name] = table_result
                validator.table_summaries.append(table_result)
            except Exception as e:
                logger.error(f"Error validating queries: {e}")
                return
        else:
            results = validator.validate_all_tables(rs_schema, sf_schema)

        # Reports
        logger.info("\nGenerating reports...")
        report_gen = ReportGenerator(validator)
        report_file = report_gen.generate_excel_report()

        # Summary to console / log
        logger.info("\n" + "=" * 60)
        logger.info("VALIDATION SUMMARY")
        logger.info("=" * 60)

        total    = len(results)
        passed   = sum(1 for r in results.values() if r.status == "PASS")
        failed   = sum(1 for r in results.values() if r.status == "FAIL")
        warnings = sum(1 for r in results.values() if r.status == "WARNING")

        logger.info(f"Total Tables : {total}")
        logger.info(f"■ Passed     : {passed}  ({passed/total*100:.1f}%)" if total else "■ Passed: 0")
        logger.info(f"■ Failed     : {failed}  ({failed/total*100:.1f}%)" if total else "■ Failed: 0")
        logger.info(f"■ Warnings   : {warnings} ({warnings/total*100:.1f}%)" if total else "■ Warnings: 0")
        logger.info("=" * 60)

        if failed:
            logger.warning("\nFailed Tables:")
            for tbl, res in results.items():
                if res.status == "FAIL":
                    logger.warning(f"  - {tbl}: {res.details}")

        if warnings:
            logger.info("\nTables with Warnings:")
            for tbl, res in results.items():
                if res.status == "WARNING":
                    logger.info(f"  - {tbl}: {res.details}")

        logger.info(f"\n■ Report saved : {report_file}")
        logger.info("■ Detailed logs: row_hash_validation.log")

    finally:
        rs_conn.disconnect()
        sf_conn.disconnect()
        logger.info("\nConnections closed.")


if __name__ == "__main__":
    main()
