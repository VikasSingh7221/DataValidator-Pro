"""
Report generation — Excel, CSV, and ZIP bundle for download.
"""

import io
import os
import zipfile
import logging
import pandas as pd
from datetime import datetime, timezone, timedelta
from typing import List, Optional

from core.models import ValidationResult, TableValidation

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError:
    Workbook = None


class ReportGenerator:
    """
    Generates Excel summary reports + per-table CSVs,
    and bundles everything into a downloadable ZIP.
    """

    # Color palette
    COLORS = {
        "PASS": "70AD47",
        "WARNING": "FFC000",
        "FAIL": "FF4444",
        "HEADER_BG": "1B2838",
        "HEADER_FG": "FFFFFF",
    }

    def __init__(
        self,
        table_summaries: List[TableValidation],
        detailed_results: List[ValidationResult],
        source_label: str = "Source",
        target_label: str = "Target",
    ):
        self.summaries = table_summaries
        self.results = detailed_results
        self.source_label = source_label
        self.target_label = target_label
        ist = timezone(timedelta(hours=5, minutes=30))
        self.timestamp = datetime.now(ist).strftime("%Y%m%d_%H%M%S")

    # ── Public: Generate ZIP bundle ───────────────────────────────────────────

    def generate_zip_bundle(self) -> bytes:
        """
        Generate a complete ZIP containing:
          - summary_report.xlsx (multi-sheet summary)
          - summary.csv
          - <table_name>/comparison.csv   (for each table)
          - <table_name>/source_sample.csv
          - <table_name>/target_sample.csv
        Returns raw bytes suitable for st.download_button.
        """
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            # 1) Summary Excel
            excel_bytes = self._generate_excel_bytes()
            zf.writestr(f"validation_report_{self.timestamp}.xlsx", excel_bytes)

            # 2) Summary CSV
            csv_bytes = self._generate_summary_csv()
            zf.writestr("summary.csv", csv_bytes)

            # 3) Per-table detail CSVs
            tables_with_results = {}
            for r in self.results:
                tables_with_results.setdefault(r.table_name, []).append(r)

            for table_name, table_results in tables_with_results.items():
                safe_name = table_name.replace("/", "_").replace("\\", "_")
                rows = []
                for r in table_results:
                    rows.append({
                        "Validation_Type": r.validation_type,
                        "Status": r.status,
                        "Source_Value": str(r.source_value),
                        "Target_Value": str(r.target_value),
                        "Difference": str(r.difference) if r.difference else "",
                        "Column": r.column_name,
                        "Details": r.details,
                    })
                df = pd.DataFrame(rows)
                csv_data = df.to_csv(index=False)
                zf.writestr(f"{safe_name}/comparison_details.csv", csv_data)

            # 4) Statistics sheet as CSV
            stats_csv = self._generate_statistics_csv()
            zf.writestr("statistics.csv", stats_csv)

        zip_buffer.seek(0)
        return zip_buffer.getvalue()

    # ── Public: Individual outputs ────────────────────────────────────────────

    def get_summary_dataframe(self) -> pd.DataFrame:
        """Summary table as a DataFrame (for Streamlit display)."""
        rows = []
        for s in self.summaries:
            rows.append({
                "Table": s.table_name,
                "Status": s.status,
                "Total Checks": s.total_checks,
                "Passed": s.passed_checks,
                "Failed": s.failed_checks,
                "Warnings": s.warning_checks,
                "Pass Rate": f"{s.pass_rate:.1f}%",
                "Source Rows": f"{s.row_count_source:,}",
                "Target Rows": f"{s.row_count_target:,}",
                "Row Diff %": f"{s.row_count_diff_pct:.2f}%",
                "Sampled": s.sampled_rows,
                "Duration (s)": f"{s.duration_seconds:.2f}",
                "Validations": ", ".join(s.validation_types_run),
            })
        return pd.DataFrame(rows)

    def get_detailed_dataframe(self) -> pd.DataFrame:
        """All detailed results as a DataFrame."""
        rows = []
        for r in self.results:
            rows.append({
                "Table": r.table_name,
                "Type": r.validation_type,
                "Status": r.status,
                "Source Value": str(r.source_value),
                "Target Value": str(r.target_value),
                "Column": r.column_name,
                "Details": r.details,
            })
        return pd.DataFrame(rows)

    def get_statistics(self) -> dict:
        """Aggregate statistics for the dashboard."""
        total = len(self.summaries)
        passed = sum(1 for s in self.summaries if s.status == "PASS")
        failed = sum(1 for s in self.summaries if s.status == "FAIL")
        warned = sum(1 for s in self.summaries if s.status == "WARNING")

        total_checks = sum(s.total_checks for s in self.summaries)
        passed_checks = sum(s.passed_checks for s in self.summaries)
        failed_checks = sum(s.failed_checks for s in self.summaries)

        total_src_rows = sum(s.row_count_source for s in self.summaries)
        total_tgt_rows = sum(s.row_count_target for s in self.summaries)
        total_duration = sum(s.duration_seconds for s in self.summaries)

        return {
            "total_tables": total,
            "passed_tables": passed,
            "failed_tables": failed,
            "warning_tables": warned,
            "pass_rate": (passed / total * 100) if total > 0 else 0,
            "total_checks": total_checks,
            "passed_checks": passed_checks,
            "failed_checks": failed_checks,
            "total_source_rows": total_src_rows,
            "total_target_rows": total_tgt_rows,
            "total_duration": round(total_duration, 2),
        }

    # ── Private: Excel generation ─────────────────────────────────────────────

    def _generate_excel_bytes(self) -> bytes:
        """Generate multi-sheet Excel workbook as bytes."""
        if Workbook is None:
            return b""

        wb = Workbook()
        wb.remove(wb.active)

        self._sheet_summary(wb)
        self._sheet_details(wb)
        self._sheet_failures(wb)
        self._sheet_statistics(wb)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def _header_style(self, ws):
        fill = PatternFill(start_color=self.COLORS["HEADER_BG"],
                           end_color=self.COLORS["HEADER_BG"], fill_type="solid")
        font = Font(bold=True, color=self.COLORS["HEADER_FG"])
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _status_fill(self, status: str) -> PatternFill:
        color = self.COLORS.get(status, "FFFFFF")
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def _sheet_summary(self, wb):
        ws = wb.create_sheet("Summary", 0)
        ws.append([
            "Table", "Status", "Total Checks", "Passed", "Failed",
            "Warnings", "Pass Rate", "Source Rows", "Target Rows",
            "Row Diff %", "Sampled", "Duration (s)", "Validations",
        ])
        self._header_style(ws)

        for s in self.summaries:
            ws.append([
                s.table_name, s.status, s.total_checks, s.passed_checks,
                s.failed_checks, s.warning_checks, f"{s.pass_rate:.1f}%",
                s.row_count_source, s.row_count_target,
                f"{s.row_count_diff_pct:.2f}%", s.sampled_rows,
                f"{s.duration_seconds:.2f}", ", ".join(s.validation_types_run),
            ])
            ws.cell(row=ws.max_row, column=2).fill = self._status_fill(s.status)

        ws.column_dimensions["A"].width = 40
        for col in "BCDEFGHIJKLM":
            ws.column_dimensions[col].width = 16

    def _sheet_details(self, wb):
        ws = wb.create_sheet("Detailed Results")
        ws.append(["Table", "Type", "Status", "Source Value", "Target Value", "Column", "Details"])
        self._header_style(ws)

        for r in self.results:
            ws.append([
                r.table_name, r.validation_type, r.status,
                str(r.source_value), str(r.target_value),
                r.column_name, r.details[:200],
            ])
            ws.cell(row=ws.max_row, column=3).fill = self._status_fill(r.status)

        ws.column_dimensions["A"].width = 40
        for col in "BCDEFG":
            ws.column_dimensions[col].width = 20

    def _sheet_failures(self, wb):
        ws = wb.create_sheet("Failures & Warnings")
        failed = [r for r in self.results if r.status == "FAIL"]
        warned = [r for r in self.results if r.status == "WARNING"]

        ws.append(["FAILURES:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="FF0000")
        ws.append(["Table", "Type", "Source", "Target", "Column", "Details"])
        self._header_style(ws)

        for r in failed:
            ws.append([
                r.table_name, r.validation_type,
                str(r.source_value), str(r.target_value),
                r.column_name, r.details[:200],
            ])

        ws.append([])
        ws.append(["WARNINGS:"])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color="FFC000")
        ws.append(["Table", "Type", "Source", "Target", "Column", "Details"])

        for r in warned:
            ws.append([
                r.table_name, r.validation_type,
                str(r.source_value), str(r.target_value),
                r.column_name, r.details[:200],
            ])

        ws.column_dimensions["A"].width = 40
        for col in "BCDEF":
            ws.column_dimensions[col].width = 20

    def _sheet_statistics(self, wb):
        ws = wb.create_sheet("Statistics")
        stats = self.get_statistics()

        def pct(n, d):
            return f"{n} ({n/d*100:.1f}%)" if d > 0 else f"{n} (N/A)"

        rows = [
            ["DATA VALIDATION SUMMARY", ""],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Source", self.source_label],
            ["Target", self.target_label],
            ["", ""],
            ["TABLE STATISTICS", ""],
            ["Total Tables", stats["total_tables"]],
            ["Passed", pct(stats["passed_tables"], stats["total_tables"])],
            ["Failed", pct(stats["failed_tables"], stats["total_tables"])],
            ["Warnings", pct(stats["warning_tables"], stats["total_tables"])],
            ["", ""],
            ["CHECK RESULTS", ""],
            ["Total Checks", stats["total_checks"]],
            ["Passed", pct(stats["passed_checks"], stats["total_checks"])],
            ["Failed", pct(stats["failed_checks"], stats["total_checks"])],
            ["", ""],
            ["DATA VOLUME", ""],
            ["Total Source Rows", f"{stats['total_source_rows']:,}"],
            ["Total Target Rows", f"{stats['total_target_rows']:,}"],
            ["Total Duration", f"{stats['total_duration']:.2f}s"],
        ]

        for row in rows:
            ws.append(row)

        for row_cells in ws.iter_rows(min_row=1, max_row=len(rows)):
            val = row_cells[0].value
            if val and isinstance(val, str) and val.isupper():
                for cell in row_cells:
                    cell.font = Font(bold=True, size=12)
                    cell.fill = PatternFill(
                        start_color="1B2838", end_color="1B2838", fill_type="solid"
                    )
                    cell.font = Font(bold=True, size=12, color="FFFFFF")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 30

    # ── Private: CSV helpers ──────────────────────────────────────────────────

    def _generate_summary_csv(self) -> str:
        return self.get_summary_dataframe().to_csv(index=False)

    def _generate_statistics_csv(self) -> str:
        stats = self.get_statistics()
        rows = [[k, v] for k, v in stats.items()]
        return pd.DataFrame(rows, columns=["Metric", "Value"]).to_csv(index=False)
